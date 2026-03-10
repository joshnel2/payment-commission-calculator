import os
import csv
import io
from datetime import datetime
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openai import AzureOpenAI

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)
app.secret_key = 'payment-commission-calculator-secret-key'

# Azure OpenAI Client
def get_azure_client():
    return AzureOpenAI(
        azure_endpoint=os.environ.get("AZURE_OPENAI_ENDPOINT"),
        api_key=os.environ.get("AZURE_OPENAI_API_KEY"),
        api_version=os.environ.get("AZURE_OPENAI_API_VERSION", "2024-02-15-preview"),
        deployment=os.environ.get("AZURE_OPENAI_DEPLOYMENT_NAME")
    )

# Neglected attorneys (skip entirely - no commission)
NEGLECTED_ATTORNEYS = {
    'dorf', 'nelson', 'zauderer', 'scarpino', 'matetsky', 'cohen', 'shehigian', 
    'lyddane', 'fraser', 'bennett', 'weissberg', 'kastner', 'taboada', 'fiorillo', 
    'kushner', 'popovic', 'post', 'yi', 'maratos', 'oropeza', 'furrer', 'choi', 
    'degennaro', 'evangelista', 'mali', 'spadinger', 'diven', 'shea', 'dacey', 'andrieux'
}

# Rate table: (own_orig, other_orig, other_work)
RATE_TABLE = {
    'alcott, daniel': (0.60, 0.40, 0.15),
    'alcott, laura': (0.60, 0.40, 0.20),
    'batsedis, olga': (0.20, 0.00, 0.10),
    'brady, bruce': (0.00, 0.00, 0.00),
    'chivily, nicholas': (0.60, 0.40, 0.20),
    'flanders, martha': (0.60, 0.40, 0.15),
    'gagas, william': (0.10, 0.00, 0.10),
    'greenberg, margery': (0.50, 0.40, 0.10),
    'hiltzik, sara': (0.20, 0.00, 0.10),
    'jacobowitz, howard': (0.50, 0.40, 0.10),
    'jahnsen, daniel': (0.50, 0.35, 0.15),
    'kaupelis, ryan': (0.50, 0.00, 0.10),
    'klein, richard': (0.50, 0.40, 0.20),
    'lavery, jason': (0.10, 0.00, 0.10),
    'locke, scott': (0.50, 0.40, 0.15),
    'makris, james': (0.125, 0.00, 0.05),
    'marks, andrew': (0.60, 0.40, 0.20),
    'nagler, vincent': (0.10, 0.00, 0.10),
    'noto, paul': (0.60, 0.40, 0.20),
    'oliver, kenneth': (0.50, 0.30, 0.15),
    'pagano, robert': (0.00, 0.00, 0.00),
    'sherwood, peter': (0.50, 0.40, 0.15),
    'stechel, ira': (0.40, 0.30, 0.15),
    'steinberg, jeffrey': (0.40, 0.40, 0.15),
    'warner, david': (0.20, 0.20, 0.00),
}

# Required CSV columns
REQUIRED_COLUMNS = [
    'Payment/Credit Note date',
    'User',
    'Matter number',
    'Originating attorney',
    'Bill issued at date',
    'Bill due at date',
    'Bill last sent at date',
    'Collected hours value'
]


def normalize_name(name, is_billing=True):
    """Normalize name based on format."""
    if not name or not isinstance(name, str):
        return ''
    
    name = name.strip()
    
    if is_billing:
        # Billing format: "Last, First" -> "last, first"
        return name.lower()
    else:
        # Originating format: "First Last" -> "last, first"
        # Check for special case: "Andrew_Old Marks_old" or "marks_old"/"andrew_old"
        name_lower = name.lower()
        if 'marks_old' in name_lower or 'andrew_old' in name_lower:
            return 'marks, andrew'
        
        # Normal conversion: "First Last" -> "last, first"
        parts = name_lower.split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        return name_lower


def is_neglected(attorney):
    """Check if attorney is in neglected list."""
    if not attorney:
        return False
    # Extract last name (first part before comma)
    last_name = attorney.split(',')[0].strip()
    return last_name in NEGLECTED_ATTORNEYS


def get_rate(attorney):
    """Get rate for attorney from rate table."""
    return RATE_TABLE.get(attorney, None)


def calculate_commission(row):
    """Calculate commission for a single row."""
    # Get original values
    payment_date = row.get('Payment/Credit Note date', '')
    billing_user = row.get('User', '')
    matter_number = row.get('Matter number', '')
    originating_attorney = row.get('Originating attorney', '')
    bill_issued_date = row.get('Bill issued at date', '')
    bill_due_date = row.get('Bill due at date', '')
    bill_last_sent_date = row.get('Bill last sent at date', '')
    collected_hours = row.get('Collected hours value', 0)
    
    # Convert collected_hours to float
    try:
        collected_hours = float(collected_hours) if collected_hours else 0
    except (ValueError, TypeError):
        collected_hours = 0
    
    # Step 1: Normalize names
    billing_user_normalized = normalize_name(billing_user, is_billing=True)
    originating_normalized = normalize_name(originating_attorney, is_billing=False)
    
    # Determine if own work
    is_own = (billing_user_normalized == originating_normalized)
    
    # Step 2: Check if billing user is neglected
    if is_neglected(billing_user_normalized):
        return {
            'payment_date': payment_date,
            'billing_user': billing_user,
            'matter_number': matter_number,
            'originating_attorney': originating_attorney,
            'bill_issued_date': bill_issued_date,
            'bill_due_date': bill_due_date,
            'bill_last_sent_date': bill_last_sent_date,
            'collected_hours': collected_hours,
            'user_pct': 0,
            'collected_user': 0,
            'collected_originator': None,
            'originator_pct': None,
            'billing_user_normalized': billing_user_normalized,
            'originating_normalized': originating_normalized
        }
    
    # Get billing user rate
    billing_rate = get_rate(billing_user_normalized)
    
    # Initialize default values
    user_pct = 0
    orig_pct = 0
    
    # Step 5: Special Exceptions (in priority order)
    
    # Exception 1 - Pagano
    if billing_user_normalized == 'pagano, robert':
        if 'axis' in matter_number.lower():
            user_pct = 0.12
        else:
            user_pct = 0
        orig_pct = 0  # Never generates originator commission
    # Exception 2 - Warner
    elif billing_user_normalized == 'warner, david' and not is_own:
        if originating_normalized == 'marks, andrew':
            user_pct = 0.30
        else:
            user_pct = 0.20
    # Exception 3 - Locke
    elif billing_user_normalized == 'locke, scott':
        matter_lower = matter_number.lower()
        if 'maesa' in matter_lower or 'grande' in matter_lower:
            user_pct = 0.50
        elif billing_rate:
            user_pct = billing_rate[1] if not is_own else billing_rate[0]
    # Exception 4 - Jahnsen originator
    elif originating_normalized == 'jahnsen, daniel' and not is_own:
        if billing_user_normalized in ('flanders, martha', 'alcott, laura'):
            user_pct = 0.50
            orig_pct = 0.15
        elif billing_rate:
            user_pct = 0.05
            orig_pct = 0.15
    # Standard calculation
    else:
        if billing_rate:
            if is_own:
                user_pct = billing_rate[0]
            else:
                user_pct = billing_rate[1]
    
    # Calculate originator commission (if not own work and not neglected)
    if not is_own and not is_neglected(originating_normalized):
        orig_rate = get_rate(originating_normalized)
        if orig_rate:
            # Use other_work rate for originator
            orig_pct = orig_rate[2]
    
    # Calculate collected amounts
    collected_user = collected_hours * user_pct
    
    # Originator collected (blank if own work)
    if is_own:
        collected_originator = None
        originator_pct = None
    else:
        collected_originator = collected_hours * orig_pct if orig_pct else 0
        originator_pct = orig_pct if orig_pct else 0
    
    return {
        'payment_date': payment_date,
        'billing_user': billing_user,
        'matter_number': matter_number,
        'originating_attorney': originating_attorney,
        'bill_issued_date': bill_issued_date,
        'bill_due_date': bill_due_date,
        'bill_last_sent_date': bill_last_sent_date,
        'collected_hours': collected_hours,
        'user_pct': user_pct,
        'collected_user': collected_user,
        'collected_originator': collected_originator,
        'originator_pct': originator_pct,
        'billing_user_normalized': billing_user_normalized,
        'originating_normalized': originating_normalized
    }


def parse_csv(file_storage):
    """Parse CSV file and return list of dictionaries."""
    stream = io.StringIO(file_storage.stream.read().decode('UTF-8'), newline=None)
    reader = csv.DictReader(stream)
    
    # Validate columns
    if not reader.fieldnames:
        raise ValueError("CSV file is empty or has no headers")
    
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in reader.fieldnames]
    if missing_cols:
        raise ValueError(f"Missing required columns: {', '.join(missing_cols)}")
    
    rows = list(reader)
    return rows


def create_excel(results):
    """Create Excel workbook with commission data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Group results by billing user
    user_data = {}
    for row in results:
        user = row['billing_user_normalized']
        if user not in user_data:
            user_data[user] = []
        user_data[user].append(row)
    
    # Sort users alphabetically by last name
    sorted_users = sorted(user_data.keys(), key=lambda x: x.split(',')[0].strip())
    
    # Determine date range
    all_dates = []
    for rows in user_data.values():
        for row in rows:
            if row['payment_date']:
                try:
                    all_dates.append(datetime.strptime(row['payment_date'], '%m/%d/%Y'))
                except:
                    try:
                        all_dates.append(datetime.strptime(row['payment_date'], '%Y-%m-%d'))
                    except:
                        pass
    
    start_date = min(all_dates).strftime('%m/%d/%Y') if all_dates else ''
    end_date = max(all_dates).strftime('%m/%d/%Y') if all_dates else ''
    
    # Accounting format
    accounting_format = '( #,##0.00_);( (#,##0.00);( "-??);(@)'
    
    # Create sheet for each user
    for user_normalized in sorted_users:
        rows_data = user_data[user_normalized]
        
        # Convert "last, first" to "First Last" for tab name
        parts = user_normalized.split(',')
        if len(parts) >= 2:
            last_name = parts[0].strip().capitalize()
            first_name = parts[1].strip().capitalize()
            sheet_name = f"{first_name} {last_name}"
        else:
            sheet_name = user_normalized.capitalize()
        
        # Limit to 31 characters
        sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Set font
        calibri_11 = Font(name='Calibri', size=11)
        
        # Row 1: Title
        ws['B1'] = f"{sheet_name} {start_date} - {end_date}"
        ws['B1'].font = calibri_11
        ws['B1'].alignment = Alignment(horizontal='left')
        
        # Row 2: Headers
        headers = [
            'Payment/Credit Note date',
            'User',
            'Matter number',
            'Originating attorney',
            'Bill issued at date',
            'Bill due at date',
            'Bill last sent at date',
            'Collected hours value',
            'user percentage',
            'collected user',
            'Collected Originator',
            'Originator Percentage'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = calibri_11
        
        # Rows 3+: Data
        for row_idx, row_data in enumerate(rows_data, start=3):
            # Columns 1-7: Original data
            ws.cell(row=row_idx, column=1, value=row_data['payment_date']).font = calibri_11
            ws.cell(row=row_idx, column=2, value=row_data['billing_user']).font = calibri_11
            ws.cell(row=row_idx, column=3, value=row_data['matter_number']).font = calibri_11
            ws.cell(row=row_idx, column=4, value=row_data['originating_attorney']).font = calibri_11
            ws.cell(row=row_idx, column=5, value=row_data['bill_issued_date']).font = calibri_11
            ws.cell(row=row_idx, column=6, value=row_data['bill_due_date']).font = calibri_11
            ws.cell(row=row_idx, column=7, value=row_data['bill_last_sent_date']).font = calibri_11
            
            # Column 8: Collected hours value (accounting format)
            cell_8 = ws.cell(row=row_idx, column=8, value=row_data['collected_hours'])
            cell_8.font = calibri_11
            cell_8.number_format = accounting_format
            
            # Column 9: user percentage
            cell_9 = ws.cell(row=row_idx, column=9, value=row_data['user_pct'])
            cell_9.font = calibri_11
            
            # Column 10: collected user (accounting format)
            cell_10 = ws.cell(row=row_idx, column=10, value=row_data['collected_user'])
            cell_10.font = calibri_11
            cell_10.number_format = accounting_format
            
            # Column 11: Collected Originator
            if row_data['collected_originator'] is not None:
                cell_11 = ws.cell(row=row_idx, column=11, value=row_data['collected_originator'])
                cell_11.font = calibri_11
                cell_11.number_format = accounting_format
            else:
                ws.cell(row=row_idx, column=11, value=None).font = calibri_11
            
            # Column 12: Originator Percentage
            if row_data['originator_pct'] is not None:
                cell_12 = ws.cell(row=row_idx, column=12, value=row_data['originator_pct'])
                cell_12.font = calibri_11
            else:
                ws.cell(row=row_idx, column=12, value=None).font = calibri_11
        
        # Final Row: TOTAL
        total_row = len(rows_data) + 3
        ws.cell(row=total_row, column=1, value=f"TOTAL FOR {sheet_name}").font = calibri_11
        
        # Column 11: SUM of all collected user values
        total_collected_user = sum(row['collected_user'] for row in rows_data)
        cell_total = ws.cell(row=total_row, column=11, value=total_collected_user)
        cell_total.font = calibri_11
        cell_total.number_format = accounting_format
        
        # Column 12: 0
        ws.cell(row=total_row, column=12, value=0).font = calibri_11
    
    return wb


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith('.csv'):
            try:
                # Parse CSV
                rows = parse_csv(file)
                
                # Process each row
                results = []
                for row in rows:
                    processed = calculate_commission(row)
                    results.append(processed)
                
                # Create Excel
                wb = create_excel(results)
                
                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                return send_file(
                    buffer,
                    as_attachment=True,
                    download_name='commission_allocation.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except ValueError as e:
                flash(str(e), 'error')
            except Exception as e:
                flash(f'Error processing file: {str(e)}', 'error')
        else:
            flash('Please upload a CSV file', 'error')
    
    return render_template('index.html')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)